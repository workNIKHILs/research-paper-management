import json
import csv
import io
from datetime import date

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse

from .models import Publication, Faculty, INDEXED_CHOICES, TYPE_CHOICES, QUARTILE_CHOICES
from .forms import LoginForm, FacultyRegisterForm, ExcelUploadForm, PublicationForm

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ──────────────────────────────────────
#  AUTH
# ──────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        login(request, form.get_user())
        messages.success(request, f"Welcome back, {form.get_user().username}!")
        return redirect('dashboard')
    return render(request, 'core/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, "Logged out successfully.")
    return redirect('login')


def register_view(request):
    form = FacultyRegisterForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        Faculty.objects.create(
            user=user,
            name=form.cleaned_data['name'],
        )
        messages.success(request, "Account created! Please login.")
        return redirect('login')
    return render(request, 'core/register.html', {'form': form})


# ──────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────

def _get_publications(request):
    """Return all publications for all authenticated users."""
    return Publication.objects.select_related('faculty', 'faculty__user').all()


# ──────────────────────────────────────
#  DASHBOARD
# ──────────────────────────────────────

@login_required
def dashboard(request):
    pubs = _get_publications(request)

    # Counts by indexed type
    scopus_count = pubs.filter(indexed='SCOPUS').count()
    wos_count = pubs.filter(indexed='WOS').count()
    wos_sci_count = pubs.filter(indexed='WOS-SCI').count()
    wos_esci_count = pubs.filter(indexed='WOS-ESCI').count()

    # Year-wise data for chart
    year_labels = list(range(2018, 2025))
    year_counts = []
    for y in year_labels:
        year_counts.append(pubs.filter(year=y).count())

    # Faculty-wise data for chart (top 15)
    faculty_data = (
        pubs.values('faculty_name')
        .annotate(count=Count('sl_no'))
        .order_by('-count')[:15]
    )
    faculty_labels = [item['faculty_name'] for item in faculty_data]
    faculty_counts = [item['count'] for item in faculty_data]

    context = {
        'total': pubs.count(),
        'scopus_count': scopus_count,
        'wos_count': wos_count,
        'wos_sci_count': wos_sci_count,
        'wos_esci_count': wos_esci_count,
        'year_labels': year_labels,
        'year_counts': year_counts,
        'faculty_labels': faculty_labels,
        'faculty_counts': faculty_counts,
        'recent': pubs.order_by('-created_at')[:10],
    }
    return render(request, 'core/dashboard.html', context)


# ──────────────────────────────────────
#  PUBLICATIONS LIST + FILTERS
# ──────────────────────────────────────

@login_required
def publication_list(request):
    pubs = _get_publications(request)
    total_count = pubs.count()

    # Apply filters
    faculty_name = request.GET.get('faculty_name', '').strip()
    title = request.GET.get('title', '').strip()
    year = request.GET.get('year', '')
    indexed_list = request.GET.getlist('indexed')
    journal = request.GET.get('journal', '').strip()
    pub_type = request.GET.get('type', '')
    quartile = request.GET.get('quartile', '')

    if faculty_name:
        pubs = pubs.filter(faculty_name__icontains=faculty_name)
    if title:
        pubs = pubs.filter(title__icontains=title)
    if year:
        pubs = pubs.filter(year=int(year))
    if indexed_list:
        pubs = pubs.filter(indexed__in=indexed_list)
    if journal:
        pubs = pubs.filter(journal_conference__icontains=journal)
    if pub_type:
        pubs = pubs.filter(type=pub_type)
    if quartile:
        pubs = pubs.filter(quartile=quartile)

    context = {
        'publications': pubs,
        'total_count': total_count,
        'filtered_count': pubs.count(),
        'filters': request.GET,
        'years': list(range(2018, 2025)),
        'indexed_choices': INDEXED_CHOICES,
        'type_choices': TYPE_CHOICES,
        'quartile_choices': QUARTILE_CHOICES,
    }
    return render(request, 'core/publications.html', context)


# ──────────────────────────────────────
#  SELF-UPLOAD (Manual Form)
# ──────────────────────────────────────

@login_required
def self_upload(request):
    """Faculty manually uploads a single publication via form."""
    # Get or create faculty for logged-in user
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        faculty = Faculty.objects.create(user=request.user, name=request.user.username)

    form = PublicationForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        pub = form.save(commit=False)
        pub.faculty = faculty
        pub.faculty_name = faculty.name
        pub.save()
        messages.success(request, "✅ Publication added successfully!")
        return redirect('publication_list')
    elif request.method == 'POST':
        messages.error(request, "❌ Please fix the errors below.")

    return render(request, 'core/self_upload.html', {'form': form, 'faculty': faculty})


# ──────────────────────────────────────
#  UPLOAD (Excel)
# ──────────────────────────────────────

@login_required
def upload_publication(request):
    form = ExcelUploadForm()
    context = {'form': form}

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid() and openpyxl:
            f = request.FILES['file']
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active

                # Limit generator strictly to avoid memory bloat
                rows = []
                for idx, r in enumerate(ws.iter_rows(values_only=True)):
                    if idx >= 1000:
                        messages.warning(request, "Only the first 1,000 rows were loaded due to size limits.")
                        break
                    rows.append(r)
                wb.close()
            except Exception as e:
                messages.error(request, f"Error reading Excel file: {str(e)}")
                return render(request, 'core/upload.html', {'form': form})

            if len(rows) < 2:
                messages.error(request, "Excel file must have at least 2 rows (headers and data).")
                return render(request, 'core/upload.html', {'form': form})

            # Find header row dynamically
            header_row_idx = 0
            best_match_count = 0
            for i, row in enumerate(rows[:5]):
                row_str = ' '.join([str(c).lower() for c in row if c is not None])
                matches = sum(1 for kw in ['title', 'faculty', 'journal', 'conference', 'year', 'doi', 'index'] if kw in row_str)
                if matches > best_match_count:
                    best_match_count = matches
                    header_row_idx = i

            headers = [str(cell) if cell is not None else f'Column {i+1}' for i, cell in enumerate(rows[header_row_idx])]
            data_rows = []
            for row in rows[header_row_idx+1:]:
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    data_rows.append([str(cell) if cell is not None else '' for cell in row])

            # Store in session
            request.session['upload_headers'] = headers
            request.session['upload_data'] = data_rows

            # Preview first 5 rows
            preview = data_rows[:5]

            context.update({
                'headers': headers,
                'preview': preview,
                'show_preview': True,
                'total_rows': len(data_rows),
                'system_fields': [
                    'faculty_name', 'title', 'type', 'journal_conference',
                    'page_no', 'vol_issue', 'publication_date', 'doi', 'indexed', 'quartile'
                ],
                'system_field_labels': {
                    'faculty_name': 'Faculty Name',
                    'title': 'Title of Paper',
                    'type': 'Publication Type',
                    'journal_conference': 'Journal/Conference',
                    'page_no': 'Page No',
                    'vol_issue': 'Vol.No/Issue No',
                    'publication_date': 'Publication Date',
                    'doi': 'DOI',
                    'indexed': 'Indexed',
                    'quartile': 'Quartile',
                },
            })
        elif not openpyxl:
            messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        context['form'] = form

    return render(request, 'core/upload.html', context)


@login_required
def upload_confirm(request):
    if request.method != 'POST':
        return redirect('upload_publication')

    headers = request.session.get('upload_headers', [])
    data_rows = request.session.get('upload_data', [])

    if not headers or not data_rows:
        messages.error(request, "No upload data found. Please upload again.")
        return redirect('upload_publication')

    # Get column mapping from POST
    mapping = {}
    system_fields = [
        'faculty_name', 'title', 'type', 'journal_conference',
        'page_no', 'vol_issue', 'publication_date', 'doi', 'indexed', 'quartile'
    ]
    for field in system_fields:
        col_idx = request.POST.get(f'map_{field}', '')
        if col_idx != '':
            mapping[field] = int(col_idx)

    # Get or create faculty for logged-in user
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        faculty = Faculty.objects.create(user=request.user, name=request.user.username)

    count = 0
    for row in data_rows:
        try:
            pub_data = {}
            for field, idx in mapping.items():
                if idx < len(row):
                    pub_data[field] = row[idx]
                else:
                    pub_data[field] = ''

            # Parse date safely and refine formatting
            pub_date_val = str(pub_data.get('publication_date', '')).strip()
            
            # Handle native Excel datetimes smartly
            if pub_date_val.endswith(' 00:00:00'):
                date_part = pub_date_val.replace(' 00:00:00', '')
                if date_part.endswith('-01-01'):
                    # e.g., 2024-01-01 -> 2024
                    pub_date_val = date_part.split('-')[0]
                elif date_part.endswith('-01'):
                    # e.g., 2024-05-01 -> May 2024
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(date_part, "%Y-%m-%d")
                        pub_date_val = dt.strftime("%B %Y")
                    except ValueError:
                        pub_date_val = date_part
                else:
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(date_part, "%Y-%m-%d")
                        pub_date_val = dt.strftime("%d/%m/%Y")
                    except ValueError:
                        pub_date_val = date_part

            # Validate indexed value
            indexed_val = str(pub_data.get('indexed', '')).strip().upper()
            valid_indexed = [c[0] for c in INDEXED_CHOICES]
            if indexed_val not in valid_indexed:
                indexed_val = 'SCOPUS'

            # Validate type value
            type_val = str(pub_data.get('type', '')).strip().lower()
            valid_types = [c[0] for c in TYPE_CHOICES]
            if type_val not in valid_types:
                type_val = 'journal'

            # Validate quartile
            quartile_val = str(pub_data.get('quartile', '')).strip().upper()
            valid_quartiles = [c[0] for c in QUARTILE_CHOICES]
            if quartile_val not in valid_quartiles:
                quartile_val = ''

            Publication.objects.create(
                faculty=faculty,
                faculty_name=pub_data.get('faculty_name', faculty.name) or faculty.name,
                title=pub_data.get('title', ''),
                type=type_val,
                journal_conference=pub_data.get('journal_conference', ''),
                page_no=pub_data.get('page_no', ''),
                vol_issue=pub_data.get('vol_issue', ''),
                publication_date=pub_date_val,
                doi=pub_data.get('doi', ''),
                indexed=indexed_val,
                quartile=quartile_val,
            )
            count += 1
        except Exception:
            continue

    # Clear session
    request.session.pop('upload_headers', None)
    request.session.pop('upload_data', None)

    messages.success(request, f"✅ Successfully imported {count} publications!")
    return redirect('publication_list')


# ──────────────────────────────────────
#  EXPORT
# ──────────────────────────────────────

@login_required
def export_page(request):
    pubs = _get_publications(request)
    total_count = pubs.count()

    # Apply filters
    year = request.GET.get('year', '')
    faculty_name = request.GET.get('faculty_name', '').strip()
    indexed_list = request.GET.getlist('indexed')
    pub_type = request.GET.get('type', '')

    if year:
        pubs = pubs.filter(year=int(year))
    if faculty_name:
        pubs = pubs.filter(faculty_name__icontains=faculty_name)
    if indexed_list:
        pubs = pubs.filter(indexed__in=indexed_list)
    if pub_type:
        pubs = pubs.filter(type=pub_type)

    context = {
        'publications': pubs,
        'total_count': total_count,
        'filtered_count': pubs.count(),
        'filters': request.GET,
        'years': list(range(2018, 2025)),
        'indexed_choices': INDEXED_CHOICES,
        'type_choices': TYPE_CHOICES,
    }
    return render(request, 'core/export.html', context)


@login_required
def export_download(request):
    pubs = _get_publications(request)

    # Apply filters
    year = request.GET.get('year', '')
    faculty_name = request.GET.get('faculty_name', '').strip()
    indexed_list = request.GET.getlist('indexed')
    pub_type = request.GET.get('type', '')

    if year:
        pubs = pubs.filter(year=int(year))
    if faculty_name:
        pubs = pubs.filter(faculty_name__icontains=faculty_name)
    if indexed_list:
        pubs = pubs.filter(indexed__in=indexed_list)
    if pub_type:
        pubs = pubs.filter(type=pub_type)

    fmt = request.GET.get('format', 'csv')
    today = date.today().isoformat()

    headers_row = [
        'SL.NO', 'Faculty Name', 'Title', 'Type',
        'Journal/Conference', 'Page No', 'Vol/Issue',
        'Publication Date', 'DOI', 'Indexed', 'Quartile'
    ]

    if fmt == 'xlsx' and openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Publications'
        ws.append(headers_row)

        for pub in pubs:
            ws.append([
                pub.sl_no, pub.faculty_name, pub.title,
                pub.get_type_display(), pub.journal_conference,
                pub.page_no, pub.vol_issue, pub.publication_date or pub.year,
                pub.doi, pub.indexed, pub.quartile,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="publications_{today}.xlsx"'
        wb.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="publications_{today}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers_row)
        for pub in pubs:
            writer.writerow([
                pub.sl_no, pub.faculty_name, pub.title,
                pub.get_type_display(), pub.journal_conference,
                pub.page_no, pub.vol_issue, pub.publication_date or pub.year,
                pub.doi, pub.indexed, pub.quartile,
            ])
        return response

# ──────────────────────────────────────
#  EDIT / DELETE
# ──────────────────────────────────────

@login_required
def edit_publication(request, pk):
    pub = get_object_or_404(Publication, pk=pk)
    
    # Check authorization
    if not request.user.is_superuser and pub.faculty.user != request.user:
        messages.error(request, "You do not have permission to edit this publication.")
        return redirect('publication_list')
        
    form = PublicationForm(request.POST or None, request.FILES or None, instance=pub)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Publication updated successfully.")
        return redirect('publication_list')
        
    return render(request, 'core/edit_publication.html', {'form': form, 'pub': pub})

@login_required
def delete_publication(request, pk):
    pub = get_object_or_404(Publication, pk=pk)
    
    # Check authorization
    if not request.user.is_superuser and pub.faculty.user != request.user:
        messages.error(request, "You do not have permission to delete this publication.")
        return redirect('publication_list')
        
    if request.method == 'POST':
        pub.delete()
        messages.success(request, "Publication deleted successfully.")
        return redirect('publication_list')
        
    # If not POST, you might want to show a confirmation page, 
    # but for simplicity, we can redirect or show an error
    return render(request, 'core/delete_publication.html', {'pub': pub})